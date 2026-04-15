"""Inspect operations: workbook-level info and sheet listing."""

from __future__ import annotations

import datetime
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple, range_boundaries


def info(file: str | Path) -> Dict[str, Any]:
    """Return workbook-level metadata: file path, sheet names, and named range names.

    Args:
        file: Path to the .xlsx file.

    Returns:
        dict with keys:
            file (str): The file path as given.
            sheets (list[str]): Sheet names in workbook order.
            named_ranges (list[str]): Named range names defined in the workbook.
    """
    path = Path(file)
    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    except Exception as exc:
        print(f"error: {exc}", file=sys.stderr)
        sys.exit(1)

    sheets: List[str] = wb.sheetnames
    named_ranges: List[str] = list(wb.defined_names)

    wb.close()

    return {
        "file": str(file),
        "sheets": sheets,
        "named_ranges": named_ranges,
    }


def list_sheets(file: str | Path) -> Dict[str, Any]:
    """Return per-sheet metadata: name, row count, column count, and visibility.

    Args:
        file: Path to the .xlsx file.

    Returns:
        dict with key:
            sheets (list[dict]): Each dict has keys name, rows, cols, visible.
    """
    path = Path(file)
    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    except Exception as exc:
        print(f"error: {exc}", file=sys.stderr)
        sys.exit(1)

    result: List[Dict[str, Any]] = []
    for ws in wb.worksheets:
        rows = ws.max_row or 0
        cols = ws.max_column or 0
        # SheetState values: "visible", "hidden", "veryHidden"
        visible = ws.sheet_state == "visible"
        result.append(
            {
                "name": ws.title,
                "rows": rows,
                "cols": cols,
                "visible": visible,
            }
        )

    wb.close()

    return {"sheets": result}


def _convert_cell(value: Any) -> Any:
    """Convert an openpyxl cell value to a JSON-serialisable Python type.

    Conversion rules:
        None         → None
        bool         → bool  (must be checked before int since bool subclasses int)
        int          → int
        float        → float
        str          → str
        datetime     → ISO-8601 string with time component (checked before date)
        date         → ISO-8601 date string (YYYY-MM-DD)
        anything else → str(value)
    """
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return value
    if isinstance(value, str):
        return value
    if isinstance(value, datetime.datetime):
        return value.strftime("%Y-%m-%dT%H:%M:%S")
    if isinstance(value, datetime.date):
        return value.strftime("%Y-%m-%d")
    return str(value)


def read_sheet(
    file: str | Path,
    sheet: str,
    range_str: Optional[str] = None,
) -> Dict[str, Any]:
    """Return a 2D list of typed cell values for a sheet or a specified sub-range.

    Args:
        file: Path to the .xlsx file.
        sheet: Name of the sheet to read.
        range_str: Optional range in A1:H50 notation. If omitted, the entire
                   used range is returned.

    Returns:
        dict with keys:
            sheet (str): The sheet name as given.
            range (str): The actual range read, in A1:XY notation.
            rows (list[list]): 2-D list; each inner list is one row, left to right.
                Cell values are int, float, str, bool, ISO-8601 str, or None.
    """
    path = Path(file)
    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    except Exception as exc:
        print(f"error: {exc}", file=sys.stderr)
        sys.exit(1)

    if sheet not in wb.sheetnames:
        print(f"error: sheet '{sheet}' not found", file=sys.stderr)
        wb.close()
        sys.exit(1)

    ws = wb[sheet]

    if range_str is not None:
        # range_boundaries returns (min_col, min_row, max_col, max_row) — all 1-based
        min_col, min_row, max_col, max_row = range_boundaries(range_str.upper())
        used_range = range_str.upper()
    else:
        min_row = 1
        max_row = ws.max_row or 0
        min_col = 1
        max_col = ws.max_column or 0
        if max_row == 0 or max_col == 0:
            wb.close()
            return {"sheet": sheet, "range": "A1:A1", "rows": []}
        used_range = f"A1:{get_column_letter(max_col)}{max_row}"

    rows: List[List[Any]] = []
    for row in ws.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
    ):
        rows.append([_convert_cell(cell.value) for cell in row])

    wb.close()

    return {
        "sheet": sheet,
        "range": used_range,
        "rows": rows,
    }


def read_table(
    file: str | Path,
    sheet: str,
    header_row: int = 1,
    range_str: Optional[str] = None,
) -> Dict[str, Any]:
    """Return a sheet as an array-of-objects keyed by the header row.

    Args:
        file: Path to the .xlsx file.
        sheet: Name of the sheet to read.
        header_row: 1-based row number to use as the header. Default 1.
        range_str: Optional range in A1:H50 notation. If omitted, the entire
                   used range is returned.

    Returns:
        dict with keys:
            sheet (str): The sheet name as given.
            range (str): The actual range read, in A1:XY notation.
            header_row (int): The 1-based row number used as headers.
            headers (list[str]): Header names; duplicate names are suffixed
                with their column letter (e.g. "Status_C").
            rows (list[dict]): Each dict maps header name to typed cell value.
                The header row itself does not appear in rows.
    """
    path = Path(file)
    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    except Exception as exc:
        print(f"error: {exc}", file=sys.stderr)
        sys.exit(1)

    if sheet not in wb.sheetnames:
        print(f"error: sheet '{sheet}' not found", file=sys.stderr)
        wb.close()
        sys.exit(1)

    ws = wb[sheet]

    if range_str is not None:
        min_col, min_row, max_col, max_row = range_boundaries(range_str.upper())
        used_range = range_str.upper()
    else:
        min_row = 1
        max_row = ws.max_row or 0
        min_col = 1
        max_col = ws.max_column or 0
        if max_row == 0 or max_col == 0:
            wb.close()
            return {
                "sheet": sheet,
                "range": "A1:A1",
                "header_row": header_row,
                "headers": [],
                "rows": [],
            }
        used_range = f"A1:{get_column_letter(max_col)}{max_row}"

    # Read all rows in the range eagerly (read_only iterators are single-pass)
    all_rows: List[Any] = []
    for row in ws.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
    ):
        all_rows.append(list(row))

    wb.close()

    # header_row is 1-based absolute; convert to 0-based index within all_rows
    header_row_index = header_row - min_row
    if header_row_index < 0 or header_row_index >= len(all_rows):
        print(
            f"error: header_row {header_row} is outside the data range",
            file=sys.stderr,
        )
        sys.exit(1)

    header_cells = all_rows[header_row_index]

    # Build header list; suffix duplicate names with their column letter
    seen: set = set()
    headers: List[str] = []
    for cell in header_cells:
        raw = _convert_cell(cell.value)
        name = str(raw) if raw is not None else ""
        if name in seen:
            col_letter = get_column_letter(cell.column)
            name = f"{name}_{col_letter}"
        else:
            seen.add(name)
        headers.append(name)

    # Build rows, skipping the header row
    rows: List[Dict[str, Any]] = []
    for i, row in enumerate(all_rows):
        if i == header_row_index:
            continue
        row_dict: Dict[str, Any] = {}
        for j, cell in enumerate(row):
            key = headers[j] if j < len(headers) else get_column_letter(min_col + j)
            row_dict[key] = _convert_cell(cell.value)
        rows.append(row_dict)

    return {
        "sheet": sheet,
        "range": used_range,
        "header_row": header_row,
        "headers": headers,
        "rows": rows,
    }


def get_cell(
    file: str | Path,
    sheet: str,
    cell: str,
) -> Dict[str, Any]:
    """Return the typed value of a single A1-addressed cell.

    Args:
        file: Path to the .xlsx file.
        sheet: Name of the sheet.
        cell: Cell address in A1 notation (e.g. "B3").

    Returns:
        dict with keys:
            sheet (str): The sheet name as given.
            cell (str): The cell address, normalised to uppercase.
            value: The typed cell value (int, float, str, bool, ISO-8601 str,
                or None).
    """
    path = Path(file)
    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    except Exception as exc:
        print(f"error: {exc}", file=sys.stderr)
        sys.exit(1)

    if sheet not in wb.sheetnames:
        print(f"error: sheet '{sheet}' not found", file=sys.stderr)
        wb.close()
        sys.exit(1)

    ws = wb[sheet]
    cell_upper = cell.upper()

    try:
        row_idx, col_idx = coordinate_to_tuple(cell_upper)
    except Exception as exc:
        print(f"error: invalid cell address '{cell}': {exc}", file=sys.stderr)
        wb.close()
        sys.exit(1)

    # iter_rows works reliably in read_only mode; fetch the single cell
    value = None
    for row in ws.iter_rows(
        min_row=row_idx,
        max_row=row_idx,
        min_col=col_idx,
        max_col=col_idx,
    ):
        if row:
            value = _convert_cell(row[0].value)

    wb.close()

    return {
        "sheet": sheet,
        "cell": cell_upper,
        "value": value,
    }
