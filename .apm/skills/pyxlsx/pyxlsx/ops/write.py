"""Write operations: cell mutations with atomic save."""

from __future__ import annotations

import os
import sys
import tempfile
from pathlib import Path
from typing import Any, Dict, Union


def _infer_type(value: str) -> Any:
    """Infer Python type from a string value.

    Rules (in order):
        value.startswith("=")  →  str (formula string, kept as-is)
        int(value) succeeds    →  int
        float(value) succeeds  →  float
        else                   →  str
    """
    if value.startswith("="):
        return value
    try:
        return int(value)
    except ValueError:
        pass
    try:
        return float(value)
    except ValueError:
        pass
    return value


def set_cell(
    file: Union[str, Path],
    sheet: str,
    cell: str,
    value: str,
) -> Dict[str, Any]:
    """Set a single cell to an inferred-type value and save atomically.

    Type inference rules (applied to the raw string *value*):
        - Starts with ``=``       → stored as a formula string
        - Parseable as ``int``    → stored as int
        - Parseable as ``float``  → stored as float
        - Otherwise               → stored as str

    The file is written to a temp file in the same directory as *file*, then
    renamed over the original with ``os.replace()`` so the write is atomic.

    Args:
        file:  Path to the .xlsx file.
        sheet: Name of the worksheet.
        cell:  Cell address in A1 notation (e.g. ``"B3"``).
        value: The string value to write (type-inferred before storing).

    Returns:
        dict with keys:
            sheet (str): The sheet name as given.
            cell (str):  The cell address, normalised to uppercase.
            value:       The stored value after type inference.
    """
    import openpyxl  # local import to keep the module importable without openpyxl installed

    path = Path(file)
    try:
        wb = openpyxl.load_workbook(str(path), data_only=False)
    except Exception as exc:
        print(f"error: {exc}", file=sys.stderr)
        sys.exit(1)

    if sheet not in wb.sheetnames:
        print(f"error: sheet '{sheet}' not found", file=sys.stderr)
        wb.close()
        sys.exit(1)

    ws = wb[sheet]
    cell_upper = cell.upper()
    typed_value = _infer_type(value)
    ws[cell_upper] = typed_value

    # Atomic save: write to a temp file in the same directory, then replace.
    dir_path = path.parent
    try:
        fd, tmp_path = tempfile.mkstemp(dir=str(dir_path), suffix=".tmp")
        os.close(fd)
        wb.save(tmp_path)
        os.replace(tmp_path, str(path))
    except Exception as exc:
        # Clean up temp file on failure.
        try:
            os.unlink(tmp_path)
        except OSError:
            pass
        print(f"error: {exc}", file=sys.stderr)
        sys.exit(1)

    return {
        "sheet": sheet,
        "cell": cell_upper,
        "value": typed_value,
    }
