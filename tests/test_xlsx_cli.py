"""CLI integration tests for pyxlsx, exercised via click's CliRunner."""
import json

import openpyxl
import pytest
from click.testing import CliRunner

from pyxlsx.cli import cli


@pytest.fixture
def runner():
    return CliRunner()


# ── Global ────────────────────────────────────────────────────────────────────


class TestGlobal:
    def test_version_output(self, runner):
        result = runner.invoke(cli, ["--version"])
        assert result.exit_code == 0
        assert "pyxlsx" in result.output

    def test_help_lists_subcommands(self, runner):
        result = runner.invoke(cli, ["--help"])
        assert result.exit_code == 0
        for cmd in ("info", "sheet", "table", "cell", "unpack", "pack"):
            assert cmd in result.output

    def test_sheet_help(self, runner):
        result = runner.invoke(cli, ["sheet", "--help"])
        assert result.exit_code == 0
        for sub in ("list", "read", "add", "delete", "rename"):
            assert sub in result.output

    def test_table_help(self, runner):
        result = runner.invoke(cli, ["table", "--help"])
        assert result.exit_code == 0
        assert "read" in result.output

    def test_cell_help(self, runner):
        result = runner.invoke(cli, ["cell", "--help"])
        assert result.exit_code == 0
        for sub in ("get", "set"):
            assert sub in result.output


# ── info ──────────────────────────────────────────────────────────────────────


class TestInfoCmd:
    def test_json_output(self, runner, minimal_xlsx):
        result = runner.invoke(cli, ["info", str(minimal_xlsx)])
        assert result.exit_code == 0
        data = json.loads(result.output)
        assert data["sheets"] == ["Data", "Summary"]
        assert "named_ranges" in data

    def test_plain_output(self, runner, minimal_xlsx):
        result = runner.invoke(cli, ["--plain", "info", str(minimal_xlsx)])
        assert result.exit_code == 0
        assert "Data" in result.output
        assert "Summary" in result.output
        assert "Named ranges:" in result.output

    def test_missing_file_exits_1(self, runner, tmp_path):
        result = runner.invoke(cli, ["info", str(tmp_path / "missing.xlsx")])
        assert result.exit_code == 1


# ── sheet list ────────────────────────────────────────────────────────────────


class TestSheetListCmd:
    def test_json_output(self, runner, minimal_xlsx):
        result = runner.invoke(cli, ["sheet", "list", str(minimal_xlsx)])
        assert result.exit_code == 0
        data = json.loads(result.output)
        names = [s["name"] for s in data["sheets"]]
        assert "Data" in names
        assert "Summary" in names

    def test_plain_output(self, runner, minimal_xlsx):
        result = runner.invoke(cli, ["--plain", "sheet", "list", str(minimal_xlsx)])
        assert result.exit_code == 0
        assert "Data" in result.output
        assert "rows" in result.output

    def test_plain_hidden_sheet_flagged(self, runner, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Visible"
        ws2 = wb.create_sheet("Hidden")
        ws2.sheet_state = "hidden"
        path = tmp_path / "hidden.xlsx"
        wb.save(str(path))

        result = runner.invoke(cli, ["--plain", "sheet", "list", str(path)])
        assert result.exit_code == 0
        assert "[hidden]" in result.output

    def test_missing_file_exits_1(self, runner, tmp_path):
        result = runner.invoke(cli, ["sheet", "list", str(tmp_path / "missing.xlsx")])
        assert result.exit_code == 1


# ── sheet read ────────────────────────────────────────────────────────────────


class TestSheetReadCmd:
    def test_json_output(self, runner, minimal_xlsx):
        result = runner.invoke(cli, ["sheet", "read", str(minimal_xlsx), "Data"])
        assert result.exit_code == 0
        data = json.loads(result.output)
        assert data["sheet"] == "Data"
        assert len(data["rows"]) == 4

    def test_with_range(self, runner, minimal_xlsx):
        result = runner.invoke(
            cli, ["sheet", "read", str(minimal_xlsx), "Data", "--range", "A1:B2"]
        )
        assert result.exit_code == 0
        data = json.loads(result.output)
        assert data["range"] == "A1:B2"
        assert len(data["rows"]) == 2

    def test_plain_output(self, runner, minimal_xlsx):
        result = runner.invoke(cli, ["--plain", "sheet", "read", str(minimal_xlsx), "Data"])
        assert result.exit_code == 0
        assert "Name" in result.output

    def test_missing_sheet_exits_1(self, runner, minimal_xlsx):
        result = runner.invoke(cli, ["sheet", "read", str(minimal_xlsx), "NoSheet"])
        assert result.exit_code == 1


# ── sheet add ─────────────────────────────────────────────────────────────────


class TestSheetAddCmd:
    def test_json_output(self, runner, minimal_xlsx):
        result = runner.invoke(cli, ["sheet", "add", str(minimal_xlsx), "NewSheet"])
        assert result.exit_code == 0
        data = json.loads(result.output)
        assert data["name"] == "NewSheet"
        assert "position" in data

    def test_with_position(self, runner, minimal_xlsx):
        result = runner.invoke(
            cli, ["sheet", "add", str(minimal_xlsx), "FirstSheet", "--position", "1"]
        )
        assert result.exit_code == 0
        data = json.loads(result.output)
        assert data["position"] == 1

    def test_plain_output(self, runner, minimal_xlsx):
        result = runner.invoke(
            cli, ["--plain", "sheet", "add", str(minimal_xlsx), "PlainSheet"]
        )
        assert result.exit_code == 0
        assert "PlainSheet" in result.output
        assert "position" in result.output

    def test_duplicate_sheet_exits_1(self, runner, minimal_xlsx):
        result = runner.invoke(cli, ["sheet", "add", str(minimal_xlsx), "Data"])
        assert result.exit_code == 1


# ── sheet delete ──────────────────────────────────────────────────────────────


class TestSheetDeleteCmd:
    def test_json_output(self, runner, minimal_xlsx):
        result = runner.invoke(cli, ["sheet", "delete", str(minimal_xlsx), "Summary"])
        assert result.exit_code == 0
        data = json.loads(result.output)
        assert data["deleted"] == "Summary"

    def test_plain_output(self, runner, minimal_xlsx):
        result = runner.invoke(
            cli, ["--plain", "sheet", "delete", str(minimal_xlsx), "Summary"]
        )
        assert result.exit_code == 0
        assert "Summary" in result.output

    def test_delete_last_sheet_exits_1(self, runner, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Only"
        path = tmp_path / "single.xlsx"
        wb.save(str(path))

        result = runner.invoke(cli, ["sheet", "delete", str(path), "Only"])
        assert result.exit_code == 1

    def test_missing_sheet_exits_1(self, runner, minimal_xlsx):
        result = runner.invoke(cli, ["sheet", "delete", str(minimal_xlsx), "NoSheet"])
        assert result.exit_code == 1


# ── sheet rename ──────────────────────────────────────────────────────────────


class TestSheetRenameCmd:
    def test_json_output(self, runner, minimal_xlsx):
        result = runner.invoke(
            cli, ["sheet", "rename", str(minimal_xlsx), "Summary", "NewSummary"]
        )
        assert result.exit_code == 0
        data = json.loads(result.output)
        assert data["old_name"] == "Summary"
        assert data["new_name"] == "NewSummary"

    def test_plain_output(self, runner, minimal_xlsx):
        result = runner.invoke(
            cli, ["--plain", "sheet", "rename", str(minimal_xlsx), "Summary", "NewSummary"]
        )
        assert result.exit_code == 0
        assert "Summary" in result.output
        assert "NewSummary" in result.output

    def test_missing_old_name_exits_1(self, runner, minimal_xlsx):
        result = runner.invoke(
            cli, ["sheet", "rename", str(minimal_xlsx), "NoSheet", "NewName"]
        )
        assert result.exit_code == 1

    def test_new_name_exists_exits_1(self, runner, minimal_xlsx):
        result = runner.invoke(
            cli, ["sheet", "rename", str(minimal_xlsx), "Data", "Summary"]
        )
        assert result.exit_code == 1


# ── table read ────────────────────────────────────────────────────────────────


class TestTableReadCmd:
    def test_json_output(self, runner, minimal_xlsx):
        result = runner.invoke(cli, ["table", "read", str(minimal_xlsx), "Data"])
        assert result.exit_code == 0
        data = json.loads(result.output)
        assert data["headers"] == ["Name", "Count", "Score", "Active", "Timestamp"]
        assert len(data["rows"]) == 3

    def test_header_row_option(self, runner, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["intro"])
        ws.append(["A", "B"])
        ws.append([1, 2])
        path = tmp_path / "hr2.xlsx"
        wb.save(str(path))

        result = runner.invoke(
            cli, ["table", "read", str(path), "Sheet1", "--header-row", "2"]
        )
        assert result.exit_code == 0
        data = json.loads(result.output)
        assert data["headers"] == ["A", "B"]

    def test_plain_output(self, runner, minimal_xlsx):
        result = runner.invoke(cli, ["--plain", "table", "read", str(minimal_xlsx), "Data"])
        assert result.exit_code == 0
        assert "Name" in result.output
        assert "---" in result.output

    def test_missing_sheet_exits_1(self, runner, minimal_xlsx):
        result = runner.invoke(cli, ["table", "read", str(minimal_xlsx), "NoSheet"])
        assert result.exit_code == 1


# ── cell get ──────────────────────────────────────────────────────────────────


class TestCellGetCmd:
    def test_json_output(self, runner, minimal_xlsx):
        result = runner.invoke(cli, ["cell", "get", str(minimal_xlsx), "Data", "A2"])
        assert result.exit_code == 0
        data = json.loads(result.output)
        assert data["value"] == "Alpha"

    def test_int_value(self, runner, minimal_xlsx):
        result = runner.invoke(cli, ["cell", "get", str(minimal_xlsx), "Data", "B2"])
        assert result.exit_code == 0
        data = json.loads(result.output)
        assert data["value"] == 1

    def test_plain_output(self, runner, minimal_xlsx):
        result = runner.invoke(
            cli, ["--plain", "cell", "get", str(minimal_xlsx), "Data", "A2"]
        )
        assert result.exit_code == 0
        assert "Alpha" in result.output

    def test_missing_sheet_exits_1(self, runner, minimal_xlsx):
        result = runner.invoke(cli, ["cell", "get", str(minimal_xlsx), "NoSheet", "A1"])
        assert result.exit_code == 1


# ── cell set ──────────────────────────────────────────────────────────────────


class TestCellSetCmd:
    def test_json_output_int(self, runner, minimal_xlsx):
        result = runner.invoke(
            cli, ["cell", "set", str(minimal_xlsx), "Data", "A5", "99"]
        )
        assert result.exit_code == 0
        data = json.loads(result.output)
        assert data["value"] == 99
        assert data["sheet"] == "Data"
        assert data["cell"] == "A5"

    def test_json_output_float(self, runner, minimal_xlsx):
        result = runner.invoke(
            cli, ["cell", "set", str(minimal_xlsx), "Data", "A5", "3.14"]
        )
        assert result.exit_code == 0
        data = json.loads(result.output)
        assert data["value"] == pytest.approx(3.14)

    def test_json_output_str(self, runner, minimal_xlsx):
        result = runner.invoke(
            cli, ["cell", "set", str(minimal_xlsx), "Data", "A5", "hello"]
        )
        assert result.exit_code == 0
        data = json.loads(result.output)
        assert data["value"] == "hello"

    def test_plain_output(self, runner, minimal_xlsx):
        result = runner.invoke(
            cli, ["--plain", "cell", "set", str(minimal_xlsx), "Data", "A5", "42"]
        )
        assert result.exit_code == 0
        assert "A5" in result.output
        assert "42" in result.output

    def test_missing_sheet_exits_1(self, runner, minimal_xlsx):
        result = runner.invoke(
            cli, ["cell", "set", str(minimal_xlsx), "NoSheet", "A1", "value"]
        )
        assert result.exit_code == 1


# ── unpack command ────────────────────────────────────────────────────────────


class TestUnpackCmd:
    def test_json_output(self, runner, minimal_xlsx, tmp_path):
        dest = tmp_path / "out"
        result = runner.invoke(cli, ["unpack", str(minimal_xlsx), str(dest)])
        assert result.exit_code == 0
        data = json.loads(result.output)
        assert data["unpacked_dir"] == str(dest)
        assert dest.is_dir()

    def test_plain_output(self, runner, minimal_xlsx, tmp_path):
        dest = tmp_path / "plain_out"
        result = runner.invoke(cli, ["--plain", "unpack", str(minimal_xlsx), str(dest)])
        assert result.exit_code == 0
        assert str(dest) in result.output

    def test_missing_file_exits_1(self, runner, tmp_path):
        result = runner.invoke(cli, ["unpack", str(tmp_path / "missing.xlsx")])
        assert result.exit_code == 1


# ── pack command ──────────────────────────────────────────────────────────────


class TestPackCmd:
    def test_json_output(self, runner, minimal_xlsx, tmp_path):
        unpacked = tmp_path / "unpacked"
        runner.invoke(cli, ["unpack", str(minimal_xlsx), str(unpacked)])
        repacked = tmp_path / "repacked.xlsx"

        result = runner.invoke(cli, ["pack", str(unpacked), str(repacked)])
        assert result.exit_code == 0
        data = json.loads(result.output)
        assert data["output_file"] == str(repacked)
        assert repacked.exists()

    def test_plain_output(self, runner, minimal_xlsx, tmp_path):
        unpacked = tmp_path / "unpacked2"
        runner.invoke(cli, ["unpack", str(minimal_xlsx), str(unpacked)])
        repacked = tmp_path / "repacked2.xlsx"

        result = runner.invoke(cli, ["--plain", "pack", str(unpacked), str(repacked)])
        assert result.exit_code == 0
        assert str(repacked) in result.output

    def test_nonexistent_dir_exits_1(self, runner, tmp_path):
        result = runner.invoke(
            cli, ["pack", str(tmp_path / "no_dir"), str(tmp_path / "out.xlsx")]
        )
        assert result.exit_code == 1
