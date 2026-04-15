"""Tests for pyxlsx.ops.write — set_cell, add_sheet, delete_sheet, rename_sheet."""
import openpyxl
import pytest

from pyxlsx.ops.inspect import get_cell
from pyxlsx.ops.write import add_sheet, delete_sheet, rename_sheet, set_cell


# ── set_cell() ────────────────────────────────────────────────────────────────


class TestSetCell:
    def test_int_inference(self, minimal_xlsx):
        result = set_cell(str(minimal_xlsx), "Data", "A5", "42")
        assert result["value"] == 42
        assert isinstance(result["value"], int)
        assert result["sheet"] == "Data"
        assert result["cell"] == "A5"

    def test_float_inference(self, minimal_xlsx):
        result = set_cell(str(minimal_xlsx), "Data", "B5", "3.14")
        assert result["value"] == pytest.approx(3.14)

    def test_str_fallback(self, minimal_xlsx):
        result = set_cell(str(minimal_xlsx), "Data", "A5", "hello")
        assert result["value"] == "hello"
        assert isinstance(result["value"], str)

    def test_formula_stored_as_str(self, minimal_xlsx):
        result = set_cell(str(minimal_xlsx), "Data", "A5", "=SUM(B2:B4)")
        assert result["value"] == "=SUM(B2:B4)"

    def test_cell_address_normalized_uppercase(self, minimal_xlsx):
        result = set_cell(str(minimal_xlsx), "Data", "a5", "test")
        assert result["cell"] == "A5"

    def test_round_trip_with_get_cell(self, minimal_xlsx):
        set_cell(str(minimal_xlsx), "Data", "A5", "99")
        result = get_cell(str(minimal_xlsx), "Data", "A5")
        assert result["value"] == 99

    def test_missing_sheet_exits_1(self, minimal_xlsx):
        with pytest.raises(SystemExit) as exc_info:
            set_cell(str(minimal_xlsx), "NoSuchSheet", "A1", "value")
        assert exc_info.value.code == 1

    def test_atomic_write_no_tmp_files_left(self, minimal_xlsx):
        dir_path = minimal_xlsx.parent
        files_before = set(dir_path.iterdir())
        set_cell(str(minimal_xlsx), "Data", "A5", "atomic_test")
        files_after = set(dir_path.iterdir())
        new_files = files_after - files_before
        tmp_files = [f for f in new_files if f.suffix == ".tmp"]
        assert tmp_files == []


# ── add_sheet() ───────────────────────────────────────────────────────────────


class TestAddSheet:
    def test_add_at_end(self, minimal_xlsx):
        result = add_sheet(str(minimal_xlsx), "NewSheet")
        assert result["name"] == "NewSheet"
        assert result["position"] == 3

        wb = openpyxl.load_workbook(str(minimal_xlsx))
        assert "NewSheet" in wb.sheetnames
        wb.close()

    def test_add_at_position_one(self, minimal_xlsx):
        result = add_sheet(str(minimal_xlsx), "First", position=1)
        assert result["position"] == 1

        wb = openpyxl.load_workbook(str(minimal_xlsx))
        assert wb.sheetnames[0] == "First"
        wb.close()

    def test_duplicate_sheet_exits_1(self, minimal_xlsx):
        with pytest.raises(SystemExit) as exc_info:
            add_sheet(str(minimal_xlsx), "Data")
        assert exc_info.value.code == 1


# ── delete_sheet() ────────────────────────────────────────────────────────────


class TestDeleteSheet:
    def test_delete_existing_sheet(self, minimal_xlsx):
        result = delete_sheet(str(minimal_xlsx), "Summary")
        assert result["deleted"] == "Summary"

        wb = openpyxl.load_workbook(str(minimal_xlsx))
        assert "Summary" not in wb.sheetnames
        wb.close()

    def test_delete_last_sheet_exits_1(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Only"
        path = tmp_path / "single.xlsx"
        wb.save(str(path))

        with pytest.raises(SystemExit) as exc_info:
            delete_sheet(str(path), "Only")
        assert exc_info.value.code == 1

    def test_delete_nonexistent_sheet_exits_1(self, minimal_xlsx):
        with pytest.raises(SystemExit) as exc_info:
            delete_sheet(str(minimal_xlsx), "NoSuchSheet")
        assert exc_info.value.code == 1


# ── rename_sheet() ────────────────────────────────────────────────────────────


class TestRenameSheet:
    def test_rename_success(self, minimal_xlsx):
        result = rename_sheet(str(minimal_xlsx), "Summary", "NewSummary")
        assert result["old_name"] == "Summary"
        assert result["new_name"] == "NewSummary"

        wb = openpyxl.load_workbook(str(minimal_xlsx))
        assert "NewSummary" in wb.sheetnames
        assert "Summary" not in wb.sheetnames
        wb.close()

    def test_rename_missing_sheet_exits_1(self, minimal_xlsx):
        with pytest.raises(SystemExit) as exc_info:
            rename_sheet(str(minimal_xlsx), "NoSuchSheet", "AnyName")
        assert exc_info.value.code == 1

    def test_rename_to_existing_name_exits_1(self, minimal_xlsx):
        with pytest.raises(SystemExit) as exc_info:
            rename_sheet(str(minimal_xlsx), "Data", "Summary")
        assert exc_info.value.code == 1
