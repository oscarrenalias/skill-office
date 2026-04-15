"""Tests for pyxlsx.ops.inspect — info, list_sheets, read_sheet, read_table, get_cell."""
import datetime

import openpyxl
import pytest

from pyxlsx.ops.inspect import get_cell, info, list_sheets, read_sheet, read_table


# ── info() ────────────────────────────────────────────────────────────────────


class TestInfo:
    def test_returns_file_and_sheets(self, minimal_xlsx):
        result = info(str(minimal_xlsx))
        assert result["file"] == str(minimal_xlsx)
        assert result["sheets"] == ["Data", "Summary"]
        assert isinstance(result["named_ranges"], list)

    def test_named_ranges_empty(self, minimal_xlsx):
        result = info(str(minimal_xlsx))
        assert result["named_ranges"] == []

    def test_named_ranges_populated(self, tmp_path):
        from openpyxl.workbook.defined_name import DefinedName

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = 42
        dn = DefinedName(name="MyRange", attr_text="Sheet1!$A$1")
        wb.defined_names["MyRange"] = dn
        path = tmp_path / "named.xlsx"
        wb.save(str(path))

        result = info(str(path))
        assert "MyRange" in result["named_ranges"]

    def test_missing_file_exits_1(self, tmp_path):
        with pytest.raises(SystemExit) as exc_info:
            info(str(tmp_path / "missing.xlsx"))
        assert exc_info.value.code == 1


# ── list_sheets() ─────────────────────────────────────────────────────────────


class TestListSheets:
    def test_returns_sheet_list(self, minimal_xlsx):
        result = list_sheets(str(minimal_xlsx))
        sheets = result["sheets"]
        assert len(sheets) == 2
        names = [s["name"] for s in sheets]
        assert "Data" in names
        assert "Summary" in names

    def test_data_sheet_dimensions(self, minimal_xlsx):
        result = list_sheets(str(minimal_xlsx))
        data_sheet = next(s for s in result["sheets"] if s["name"] == "Data")
        assert data_sheet["rows"] == 4
        assert data_sheet["cols"] == 5

    def test_all_sheets_visible(self, minimal_xlsx):
        result = list_sheets(str(minimal_xlsx))
        for s in result["sheets"]:
            assert s["visible"] is True

    def test_hidden_sheet_flagged(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Visible"
        ws2 = wb.create_sheet("Hidden")
        ws2.sheet_state = "hidden"
        path = tmp_path / "hidden.xlsx"
        wb.save(str(path))

        result = list_sheets(str(path))
        sheet_map = {s["name"]: s for s in result["sheets"]}
        assert sheet_map["Visible"]["visible"] is True
        assert sheet_map["Hidden"]["visible"] is False

    def test_missing_file_exits_1(self, tmp_path):
        with pytest.raises(SystemExit) as exc_info:
            list_sheets(str(tmp_path / "missing.xlsx"))
        assert exc_info.value.code == 1


# ── read_sheet() ──────────────────────────────────────────────────────────────


class TestReadSheet:
    def test_full_range_row_count(self, minimal_xlsx):
        result = read_sheet(str(minimal_xlsx), "Data")
        assert result["sheet"] == "Data"
        assert result["range"] == "A1:E4"
        assert len(result["rows"]) == 4

    def test_header_row_values(self, minimal_xlsx):
        result = read_sheet(str(minimal_xlsx), "Data")
        assert result["rows"][0] == ["Name", "Count", "Score", "Active", "Timestamp"]

    def test_type_str(self, minimal_xlsx):
        rows = read_sheet(str(minimal_xlsx), "Data")["rows"]
        assert rows[1][0] == "Alpha"

    def test_type_int(self, minimal_xlsx):
        rows = read_sheet(str(minimal_xlsx), "Data")["rows"]
        assert rows[1][1] == 1
        assert isinstance(rows[1][1], int)

    def test_type_float(self, minimal_xlsx):
        rows = read_sheet(str(minimal_xlsx), "Data")["rows"]
        assert rows[1][2] == pytest.approx(1.1)

    def test_type_bool(self, minimal_xlsx):
        rows = read_sheet(str(minimal_xlsx), "Data")["rows"]
        assert rows[1][3] is True

    def test_type_datetime_iso(self, minimal_xlsx):
        rows = read_sheet(str(minimal_xlsx), "Data")["rows"]
        assert rows[1][4] == "2025-01-06T09:00:00"

    def test_subrange(self, minimal_xlsx):
        result = read_sheet(str(minimal_xlsx), "Data", range_str="A1:B2")
        assert result["range"] == "A1:B2"
        assert len(result["rows"]) == 2
        assert result["rows"][0] == ["Name", "Count"]

    def test_missing_sheet_exits_1(self, minimal_xlsx):
        with pytest.raises(SystemExit) as exc_info:
            read_sheet(str(minimal_xlsx), "NoSuchSheet")
        assert exc_info.value.code == 1

    def test_empty_sheet_returns_empty_rows(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Empty"
        path = tmp_path / "empty.xlsx"
        wb.save(str(path))

        result = read_sheet(str(path), "Empty")
        assert result["rows"] == []
        assert result["range"] == "A1:A1"

    def test_none_cell_value(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "Hello"
        ws["A2"] = None
        path = tmp_path / "none_cell.xlsx"
        wb.save(str(path))

        result = read_sheet(str(path), "Sheet1")
        assert result["rows"][1][0] is None


# ── read_table() ──────────────────────────────────────────────────────────────


class TestReadTable:
    def test_normal_table_headers(self, minimal_xlsx):
        result = read_table(str(minimal_xlsx), "Data")
        assert result["sheet"] == "Data"
        assert result["header_row"] == 1
        assert result["headers"] == ["Name", "Count", "Score", "Active", "Timestamp"]

    def test_normal_table_row_count(self, minimal_xlsx):
        result = read_table(str(minimal_xlsx), "Data")
        assert len(result["rows"]) == 3

    def test_normal_table_row_values(self, minimal_xlsx):
        rows = read_table(str(minimal_xlsx), "Data")["rows"]
        assert rows[0]["Name"] == "Alpha"
        assert rows[0]["Count"] == 1

    def test_duplicate_headers_suffixed(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["Name", "Status", "Status"])
        ws.append(["Alpha", "A", "B"])
        path = tmp_path / "dups.xlsx"
        wb.save(str(path))

        result = read_table(str(path), "Sheet1")
        headers = result["headers"]
        assert headers[0] == "Name"
        assert headers[1] == "Status"
        assert headers[2] == "Status_C"

    def test_header_row_not_one(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["intro row"])
        ws.append(["Name", "Value"])
        ws.append(["Alice", 100])
        path = tmp_path / "header2.xlsx"
        wb.save(str(path))

        result = read_table(str(path), "Sheet1", header_row=2)
        assert result["header_row"] == 2
        assert result["headers"] == ["Name", "Value"]
        data_rows = [r for r in result["rows"] if r.get("Name") == "Alice"]
        assert data_rows[0]["Value"] == 100

    def test_missing_sheet_exits_1(self, minimal_xlsx):
        with pytest.raises(SystemExit) as exc_info:
            read_table(str(minimal_xlsx), "NoSuchSheet")
        assert exc_info.value.code == 1

    def test_header_only_sheet_returns_empty_rows(self, tmp_path):
        """A sheet with only a header row returns populated headers and rows=[]."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "HeaderOnly"
        ws.append(["A", "B", "C"])
        path = tmp_path / "header_only.xlsx"
        wb.save(str(path))

        result = read_table(str(path), "HeaderOnly")
        assert result["headers"] == ["A", "B", "C"]
        assert result["rows"] == []

    def test_blank_worksheet_exits_1(self, tmp_path):
        """Known limitation: openpyxl sets max_row=1 for untouched sheets but
        iter_rows returns nothing; header-row validation fails with exit code 1."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Blank"
        path = tmp_path / "blank.xlsx"
        wb.save(str(path))

        with pytest.raises(SystemExit) as exc_info:
            read_table(str(path), "Blank")
        assert exc_info.value.code == 1

    def test_range_subset(self, minimal_xlsx):
        result = read_table(str(minimal_xlsx), "Data", range_str="A1:C4")
        assert result["headers"] == ["Name", "Count", "Score"]
        assert len(result["rows"]) == 3


# ── get_cell() ────────────────────────────────────────────────────────────────


class TestGetCell:
    def test_string_value(self, minimal_xlsx):
        result = get_cell(str(minimal_xlsx), "Data", "A2")
        assert result["sheet"] == "Data"
        assert result["cell"] == "A2"
        assert result["value"] == "Alpha"

    def test_int_value(self, minimal_xlsx):
        result = get_cell(str(minimal_xlsx), "Data", "B2")
        assert result["value"] == 1
        assert isinstance(result["value"], int)

    def test_float_value(self, minimal_xlsx):
        result = get_cell(str(minimal_xlsx), "Data", "C2")
        assert result["value"] == pytest.approx(1.1)

    def test_bool_value(self, minimal_xlsx):
        result = get_cell(str(minimal_xlsx), "Data", "D2")
        assert result["value"] is True

    def test_datetime_value(self, minimal_xlsx):
        result = get_cell(str(minimal_xlsx), "Data", "E2")
        assert result["value"] == "2025-01-06T09:00:00"

    def test_out_of_data_cell_returns_none(self, minimal_xlsx):
        result = get_cell(str(minimal_xlsx), "Data", "Z100")
        assert result["value"] is None

    def test_cell_address_normalized_uppercase(self, minimal_xlsx):
        result = get_cell(str(minimal_xlsx), "Data", "a1")
        assert result["cell"] == "A1"

    def test_missing_sheet_exits_1(self, minimal_xlsx):
        with pytest.raises(SystemExit) as exc_info:
            get_cell(str(minimal_xlsx), "NoSuchSheet", "A1")
        assert exc_info.value.code == 1

    def test_invalid_address_exits_1(self, minimal_xlsx):
        with pytest.raises(SystemExit) as exc_info:
            get_cell(str(minimal_xlsx), "Data", "NOTACELL")
        assert exc_info.value.code == 1
